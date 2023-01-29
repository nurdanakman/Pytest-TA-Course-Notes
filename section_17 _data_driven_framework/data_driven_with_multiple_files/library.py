import json
import requests
import openpyxl

class Common:

    def __init__(self, fileName, sheetName):
        global vk
        global sh
        vk = openpyxl.load_workbook(fileName, "r")
        sh = vk[sheetName]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range(1, c+1):
            cell = sh.cell (row=i, colum=1)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row = rowNumber, column=i)
            jsonRequest[keyList[i-1]]= cell.value
        return jsonRequest
