from pathlib import Path
import requests
import json
import openpyxl

def test_add_multiple_students():
    url = "https://thetestingworldapi.com/api/studentsDetails"

    vk = openpyxl.load_workbook(Path(__file__).parent / "new_student_data.xlsx", "r")
    sh = vk["Sheet1"]
    rows = sh.max_row

    f = open(Path(__file__).parent / "new_student_data.json", "r")
    json_input = f.read()
    payload = json.loads(json_input)

    for i in range(2, 20):
        cell_first_name = sh.cell(row=i, column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_birth_date = sh.cell(row=i, column=4)
        payload["first_name"]=cell_first_name.value
        payload["middle_name"]=cell_middle_name.value
        payload["last_name"]=cell_last_name.value
        payload["date_of_birth"]=cell_birth_date.value

        response = requests.post(url, payload)
        print(response.status_code)
        print(response.json())
        assert response.status_code == 201
